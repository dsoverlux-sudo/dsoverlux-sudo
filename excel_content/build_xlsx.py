# -*- coding: utf-8 -*-
"""
컴활 실무 실수 교정 워크북 생성기
대상: 컴활은 합격했으나 실무에서 반복 실수가 나오는 사람
구성: 안내 · 자가진단 · 실수교정 TOP7 · 실습(절대참조/텍스트숫자/VLOOKUP)
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "컴활실무_실수교정_v1.xlsx")

F = "맑은 고딕"
NAVY = "0F2557"; BLUE = "2563EB"; SKY = "E8EFFC"; LGRAY = "EEF0F4"
RED = "D93A3A"; GREEN = "1E8E5A"; WHITE = "FFFFFF"; GRAY = "5B6472"

def font(sz=10, b=False, color="000000"): return Font(name=F, size=sz, bold=b, color=color)
def fill(c): return PatternFill("solid", fgColor=c)
thin = Side(style="thin", color="C9CED6")
box = Border(left=thin, right=thin, top=thin, bottom=thin)
CEN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

wb = Workbook()

def title_bar(ws, text, sub, span="A1:F1"):
    c1, c2 = span.split(":")
    ws.merge_cells(span)
    cell = ws[c1]
    cell.value = text
    cell.font = font(18, True, WHITE); cell.fill = fill(NAVY); cell.alignment = LEFT
    row = int("".join(ch for ch in c1 if ch.isdigit()))
    ws.row_dimensions[row].height = 34
    sr = row + 1
    sspan = f"A{sr}:{''.join(ch for ch in c2 if ch.isalpha())}{sr}"
    ws.merge_cells(sspan)
    ws[f"A{sr}"].value = sub
    ws[f"A{sr}"].font = font(10, False, GRAY); ws[f"A{sr}"].alignment = LEFT
    ws.row_dimensions[sr].height = 20

# ───────────────────────── 1. 안내
ws = wb.active; ws.title = "안내"
for col, w in zip("ABCDEF", [3, 20, 20, 20, 16, 16]): ws.column_dimensions[col].width = w
title_bar(ws, "컴활은 합격했는데, 실무가 안 되는 이유", "자격증은 '기능을 안다'를 증명할 뿐 — 실무는 '데이터를 다룬다'입니다. 그 간격을 메우는 진단·교정 워크북.")
rows = [
    ("", ""),
    ("이 파일은 이런 분을 위한 겁니다", ""),
    ("• 컴활 1급/2급은 땄는데 회사에서 준 자료 앞에서 손이 멈춘다", ""),
    ("• VLOOKUP·SUM은 아는데 왜 틀리는지 모른 채 같은 실수가 반복된다", ""),
    ("• 함수를 '외웠지만' 언제 어디에 쓸지는 배운 적이 없다", ""),
    ("", ""),
    ("쓰는 순서", ""),
    ("① [자가진단] 시트에서 내 상태를 점수로 확인한다", ""),
    ("② [실수교정 TOP7]에서 내가 겪는 증상을 찾는다", ""),
    ("③ [실습] 시트에서 잘못된 수식 vs 올바른 수식을 직접 비교한다", ""),
    ("", ""),
    ("핵심 관점", ""),
    ("실무 엑셀의 90%는 함수 실력이 아니라 '데이터를 표(정형)로 두는 습관'에서 갈립니다.", ""),
    ("함수 암기를 늘리기 전에, 반복되는 실수 3개를 먼저 없애는 게 훨씬 빠릅니다.", ""),
]
r = 4
for a, b in rows:
    ws[f"B{r}"] = a
    if a in ("이 파일은 이런 분을 위한 겁니다", "쓰는 순서", "핵심 관점"):
        ws[f"B{r}"].font = font(12, True, NAVY)
    else:
        ws[f"B{r}"].font = font(10, False, "222222")
    ws[f"B{r}"].alignment = LEFT
    r += 1
# CTA
ws.merge_cells(f"B{r+1}:F{r+2}")
ws[f"B{r+1}"] = ("더 궁금한 실무 실수가 있으면 영상 댓글에 \"엑셀\"이라고 남겨주세요.\n"
                 "실수별 교정 영상과 실습 파일을 순서대로 올립니다.")
ws[f"B{r+1}"].font = font(11, True, WHITE); ws[f"B{r+1}"].fill = fill(BLUE)
ws[f"B{r+1}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws.row_dimensions[r+1].height = 40

# ───────────────────────── 2. 자가진단
ws = wb.create_sheet("자가진단")
for col, w in zip("ABCDEF", [3, 58, 12, 4, 4, 4]): ws.column_dimensions[col].width = w
title_bar(ws, "실무 자가진단 10문항", "각 문항에 예/아니오를 고르면 아래에서 점수와 진단이 자동으로 나옵니다. ('예'가 많을수록 재학습 필요)")
qs = [
    "수식을 아래로 복사하면 값이 틀리거나 0이 나온 적이 있다",
    "VLOOKUP이 #N/A거나 엉뚱한 값을 낸 적이 있다",
    "숫자를 SUM했는데 일부가 빠지거나 0으로 합산됐다",
    "병합된 셀 때문에 정렬·필터가 막힌 적이 있다",
    "날짜가 계산·정렬이 안 되고 글자처럼 취급된 적이 있다",
    "피벗테이블을 만들려다 원본이 정리가 안 돼 실패했다",
    "함수 이름은 아는데 실무에서 언제 쓸지 몰라 멈춘다",
    "남이 만든 수식을 열면 구조가 파악되지 않는다",
    "조건부서식·이름정의를 배웠지만 실무에서 안 쓴다",
    "자료를 받으면 정리 없이 그대로 쓴다(표로 만들지 않는다)",
]
hr = 4
ws[f"B{hr}"] = "문항"; ws[f"C{hr}"] = "예/아니오"
for c in ("B", "C"):
    ws[f"{c}{hr}"].font = font(10, True, WHITE); ws[f"{c}{hr}"].fill = fill(NAVY)
    ws[f"{c}{hr}"].alignment = CEN; ws[f"{c}{hr}"].border = box
dv = DataValidation(type="list", formula1='"예,아니오"', allow_blank=False)
ws.add_data_validation(dv)
first = hr + 1
for i, q in enumerate(qs):
    rr = first + i
    ws[f"B{rr}"] = f"{i+1}. {q}"; ws[f"B{rr}"].font = font(10); ws[f"B{rr}"].alignment = LEFT; ws[f"B{rr}"].border = box
    ws[f"C{rr}"] = "아니오"; ws[f"C{rr}"].font = font(10, True, BLUE); ws[f"C{rr}"].alignment = CEN; ws[f"C{rr}"].border = box
    ws[f"C{rr}"].fill = fill("FFFDE7")
    dv.add(ws[f"C{rr}"])
last = first + len(qs) - 1
sr = last + 2
ws[f"B{sr}"] = "'예' 개수 (점수)"; ws[f"B{sr}"].font = font(11, True, NAVY); ws[f"B{sr}"].alignment = LEFT
ws[f"C{sr}"] = f'=COUNTIF(C{first}:C{last},"예")'; ws[f"C{sr}"].font = font(12, True, RED); ws[f"C{sr}"].alignment = CEN; ws[f"C{sr}"].border = box
dr = sr + 1
ws.merge_cells(f"B{dr}:C{dr+1}")
ws[f"B{dr}"] = (f'=IF(C{sr}>=7,"진단: 재학습 단계 — 함수 암기보다 데이터 구조부터 다시. 반복 실수 3개를 먼저 없애세요.",'
               f'IF(C{sr}>=4,"진단: 실무 전환 단계 — 기초는 있습니다. 절대참조·텍스트숫자·VLOOKUP 3개만 잡으면 확 늘어요.",'
               f'"진단: 실무 감각 양호 — 이제 함수 암기가 아니라 자동화·분석으로 확장할 단계입니다."))')
ws[f"B{dr}"].font = font(11, True, GREEN); ws[f"B{dr}"].alignment = LEFT; ws[f"B{dr}"].fill = fill(SKY)
ws.row_dimensions[dr].height = 24; ws.row_dimensions[dr+1].height = 24

# ───────────────────────── 3. 실수교정 TOP7
ws = wb.create_sheet("실수교정 TOP7")
for col, w in zip("ABCDE", [3, 22, 30, 34, 20]): ws.column_dimensions[col].width = w
title_bar(ws, "실무에서 반복되는 엑셀 실수 TOP 7", "컴활 합격자에게 가장 자주 나오는 실수와 교정법. 표시된 실습 시트에서 직접 확인하세요.", "A1:E1")
hdr = ["증상 (이런 일이 난다)", "원인", "교정 (이렇게 바꾼다)", "실습"]
hr = 4
for j, h in enumerate(hdr):
    cell = ws.cell(row=hr, column=2 + j, value=h)
    cell.font = font(10, True, WHITE); cell.fill = fill(NAVY); cell.alignment = CEN; cell.border = box
data = [
    ("수식 복사하면 값이 밀려 틀린다", "상대참조만 씀 ($ 없음)", "고정할 셀은 F4로 절대참조 $A$1", "실습1"),
    ("SUM 결과가 실제보다 작다", "숫자가 '텍스트'로 저장됨", "VALUE()로 변환 후 합산·왼쪽 위 초록삼각 확인", "실습2"),
    ("VLOOKUP이 엉뚱한 값을 낸다", "4번째 인수 생략(근사일치)", "마지막에 FALSE(정확히 일치) 지정", "실습3"),
    ("정렬·필터·피벗이 막힌다", "셀 병합 남발", "병합 대신 '선택 영역 가운데 맞춤'", "—"),
    ("날짜 계산·정렬이 안 된다", "날짜가 텍스트로 입력됨", "DATEVALUE()·서식을 날짜로 통일", "—"),
    ("피벗을 못 만든다", "원본이 표(정형)가 아님", "머리글 1줄·빈행 없음·1열 1속성으로 정리", "—"),
    ("남의 수식이 안 읽힌다", "이름정의·구조 없이 좌표만", "이름정의로 =단가*수량처럼 읽히게", "—"),
]
for i, row in enumerate(data):
    rr = hr + 1 + i
    fillc = LGRAY if i % 2 else WHITE
    for j, val in enumerate(row):
        cell = ws.cell(row=rr, column=2 + j, value=val)
        cell.font = font(9.5, j == 2, NAVY if j == 2 else "222222")
        cell.alignment = LEFT if j < 3 else CEN
        cell.fill = fill(fillc); cell.border = box
    ws.row_dimensions[rr].height = 30

# ───────────────────────── 4. 실습1 절대참조
ws = wb.create_sheet("실습1_절대참조")
for col, w in zip("ABCDEFGH", [3, 12, 10, 10, 12, 16, 16, 10]): ws.column_dimensions[col].width = w
title_bar(ws, "실습 1 — 절대참조 ($) 하나로 갈리는 실무", "부가세율은 셀 1곳(H3)에만 있습니다. 복사할 때 F4로 고정하지 않으면 참조가 아래로 밀립니다.", "A1:H1")
ws.merge_cells("F3:G3")
ws["F3"] = "부가세율 →"; ws["F3"].font = font(10, True, NAVY); ws["F3"].alignment = Alignment(horizontal="right", vertical="center")
ws["H3"] = 0.1; ws["H3"].number_format = "0%"; ws["H3"].font = font(11, True, BLUE); ws["H3"].alignment = CEN; ws["H3"].border = box; ws["H3"].fill = fill("FFFDE7")
hr = 5
heads = ["상품", "수량", "단가", "금액", "❌ 부가세(상대참조)", "✅ 부가세(절대참조)"]
for j, h in enumerate(heads):
    cell = ws.cell(row=hr, column=2 + j, value=h)
    cell.font = font(9.5, True, WHITE)
    cell.fill = fill(RED if j == 4 else (GREEN if j == 5 else NAVY))
    cell.alignment = CEN; cell.border = box
items = [("A4용지", 3, 4500), ("토너", 2, 39000), ("USB", 5, 12000), ("마우스", 4, 9800)]
first = hr + 1
for i, (nm, qty, price) in enumerate(items):
    rr = first + i
    ws.cell(row=rr, column=2, value=nm).alignment = CEN
    ws.cell(row=rr, column=3, value=qty).alignment = CEN
    ws.cell(row=rr, column=4, value=price).number_format = "#,##0"
    ws.cell(row=rr, column=5, value=f"=C{rr}*D{rr}").number_format = "#,##0"
    # ❌ 상대참조: 첫 행은 H3을 가리키지만 복사되며 H4,H5.. (빈 셀) 로 밀림
    ws.cell(row=rr, column=6, value=f"=E{rr}*H{rr-3}").number_format = "#,##0"
    # ✅ 절대참조
    ws.cell(row=rr, column=7, value=f"=E{rr}*$H$3").number_format = "#,##0"
    for col in range(2, 8):
        ws.cell(row=rr, column=col).border = box
        ws.cell(row=rr, column=col).font = font(10)
note = first + len(items) + 1
ws.merge_cells(f"B{note}:G{note+1}")
ws[f"B{note}"] = ("설명:  ❌열은 =금액*H3 을 그대로 아래로 복사 — 참조가 H3→H4→H5로 밀려 빈 셀을 곱하니 둘째 줄부터 0이 됩니다.\n"
                  "        ✅열은 =금액*$H$3 — F4로 $를 붙여 고정했기에 어디로 복사해도 항상 부가세율(H3)을 봅니다.")
ws[f"B{note}"].font = font(10, False, GRAY); ws[f"B{note}"].alignment = LEFT; ws[f"B{note}"].fill = fill(LGRAY)
ws.row_dimensions[note].height = 26; ws.row_dimensions[note+1].height = 26

# ───────────────────────── 5. 실습2 텍스트숫자
ws = wb.create_sheet("실습2_텍스트숫자")
for col, w in zip("ABCDE", [3, 16, 16, 16, 20]): ws.column_dimensions[col].width = w
title_bar(ws, "실습 2 — '텍스트로 저장된 숫자'가 합계를 갉아먹는다", "다른 시스템에서 받은 자료는 숫자처럼 보여도 텍스트인 경우가 많습니다. SUM이 이를 건너뜁니다.", "A1:E1")
hr = 4
for j, h in enumerate(["항목", "값(그대로)", "값→숫자(VALUE)"]):
    cell = ws.cell(row=hr, column=2 + j, value=h)
    cell.font = font(9.5, True, WHITE); cell.fill = fill(NAVY); cell.alignment = CEN; cell.border = box
vals = [("1월", "120000"), ("2월", "98000"), ("3월", "156000"), ("4월", "143000")]
first = hr + 1
for i, (m, v) in enumerate(vals):
    rr = first + i
    ws.cell(row=rr, column=2, value=m).alignment = CEN
    ws.cell(row=rr, column=3, value=v)  # 문자열 → 텍스트 숫자
    ws.cell(row=rr, column=4, value=f"=VALUE(C{rr})").number_format = "#,##0"
    for col in range(2, 5):
        ws.cell(row=rr, column=col).border = box; ws.cell(row=rr, column=col).font = font(10)
last = first + len(vals) - 1
tr = last + 1
ws.cell(row=tr, column=2, value="합계").font = font(10, True, NAVY)
ws.cell(row=tr, column=2).alignment = CEN; ws.cell(row=tr, column=2).border = box
ws.cell(row=tr, column=3, value=f"=SUM(C{first}:C{last})").number_format = "#,##0"
ws.cell(row=tr, column=3).font = font(11, True, RED); ws.cell(row=tr, column=3).border = box; ws.cell(row=tr, column=3).alignment = CEN
ws.cell(row=tr, column=4, value=f"=SUM(D{first}:D{last})").number_format = "#,##0"
ws.cell(row=tr, column=4).font = font(11, True, GREEN); ws.cell(row=tr, column=4).border = box; ws.cell(row=tr, column=4).alignment = CEN
note = tr + 2
ws.merge_cells(f"B{note}:E{note+1}")
ws[f"B{note}"] = ("설명:  가운데 열은 텍스트 숫자라 =SUM()이 0(또는 일부만)으로 계산됩니다 — 실무에서 '합계가 안 맞아요'의 주범.\n"
                  "        오른쪽 열처럼 VALUE()로 숫자화하면 정상 합산. 셀 왼쪽 위 초록 삼각형이 '텍스트 숫자' 경고입니다.")
ws[f"B{note}"].font = font(10, False, GRAY); ws[f"B{note}"].alignment = LEFT; ws[f"B{note}"].fill = fill(LGRAY)
ws.row_dimensions[note].height = 26; ws.row_dimensions[note+1].height = 26

# ───────────────────────── 6. 실습3 VLOOKUP
ws = wb.create_sheet("실습3_VLOOKUP")
for col, w in zip("ABCDEFG", [3, 12, 14, 12, 4, 14, 16]): ws.column_dimensions[col].width = w
title_bar(ws, "실습 3 — VLOOKUP 4번째 인수(FALSE) 하나의 차이", "없는 값을 조회할 때, 근사일치(생략)는 남의 값을 조용히 돌려줍니다.", "A1:G1")
# 참조표 (사번 오름차순)
hr = 4
for j, h in enumerate(["사번", "이름", "부서"]):
    cell = ws.cell(row=hr, column=2 + j, value=h)
    cell.font = font(9.5, True, WHITE); cell.fill = fill(NAVY); cell.alignment = CEN; cell.border = box
table = [(1007, "박민준", "개발"), (1021, "최유나", "회계"), (1039, "이지후", "인사"), (1052, "김서연", "영업")]
first = hr + 1
for i, (sabun, nm, dept) in enumerate(table):
    rr = first + i
    for j, v in enumerate((sabun, nm, dept)):
        cell = ws.cell(row=rr, column=2 + j, value=v)
        cell.font = font(10); cell.alignment = CEN; cell.border = box
last = first + len(table) - 1
# 조회
qr = last + 2
ws.cell(row=qr, column=2, value="조회할 사번").font = font(10, True, NAVY)
ws.cell(row=qr, column=2).alignment = CEN
ws.cell(row=qr, column=3, value=1040).font = font(11, True, BLUE)
ws.cell(row=qr, column=3).alignment = CEN; ws.cell(row=qr, column=3).fill = fill("FFFDE7"); ws.cell(row=qr, column=3).border = box
r1 = qr + 1
ws.cell(row=r1, column=2, value="❌ 근사일치(생략)").font = font(10, True, RED)
ws.cell(row=r1, column=2).alignment = LEFT
ws.cell(row=r1, column=3, value=f"=VLOOKUP(C{qr},B{first}:D{last},2)").font = font(11, True, RED)
ws.cell(row=r1, column=3).alignment = CEN; ws.cell(row=r1, column=3).border = box
r2 = qr + 2
ws.cell(row=r2, column=2, value="✅ 정확히 일치(FALSE)").font = font(10, True, GREEN)
ws.cell(row=r2, column=2).alignment = LEFT
ws.cell(row=r2, column=3, value=f'=IFERROR(VLOOKUP(C{qr},B{first}:D{last},2,FALSE),"해당 사번 없음")').font = font(11, True, GREEN)
ws.cell(row=r2, column=3).alignment = CEN; ws.cell(row=r2, column=3).border = box
note = r2 + 2
ws.merge_cells(f"B{note}:G{note+1}")
ws[f"B{note}"] = ("설명:  1040은 실제로 없는 사번(오타)입니다. ❌ 근사일치는 없는 사번인데도 가장 가까운 이지후(1039)를 조용히 반환합니다.\n"
                  "        ✅ 정확히 일치(FALSE)는 없으면 '해당 사번 없음'이라고 알려줍니다. 실무 VLOOKUP은 거의 항상 FALSE로 씁니다.")
ws[f"B{note}"].font = font(10, False, GRAY); ws[f"B{note}"].alignment = LEFT; ws[f"B{note}"].fill = fill(LGRAY)
ws.row_dimensions[note].height = 26; ws.row_dimensions[note+1].height = 26

for s in wb.sheetnames:
    wb[s].sheet_view.showGridLines = False

wb.save(OUT)
print("SAVED", OUT)
